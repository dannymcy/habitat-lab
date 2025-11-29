import habitat_sim
import magnum as mn
import warnings
from habitat.tasks.rearrange.rearrange_sim import RearrangeSim
warnings.filterwarnings('ignore')
from habitat_sim.utils.settings import make_cfg
from matplotlib import pyplot as plt
from habitat_sim.utils import viz_utils as vut
from omegaconf import DictConfig
import numpy as np
from habitat.articulated_agents.robots import FetchRobot
from habitat.config.default import get_agent_config
from habitat.config.default_structured_configs import ThirdRGBSensorConfig, HeadRGBSensorConfig, HeadPanopticSensorConfig, TopRGBSensorConfig
from habitat.config.default_structured_configs import SimulatorConfig, HabitatSimV0Config, AgentConfig
from habitat.config.default import get_agent_config
import habitat
from habitat_sim.physics import JointMotorSettings, MotionType
from omegaconf import OmegaConf
from habitat.articulated_agent_controllers import (
    HumanoidRearrangeController,
    HumanoidSeqPoseController,
)

from habitat.config.default_structured_configs import HumanoidJointActionConfig, HumanoidPickActionConfig
from habitat.config.default_structured_configs import TaskConfig, EnvironmentConfig, DatasetConfig, HabitatConfig
from habitat.config.default_structured_configs import ArmActionConfig, BaseVelocityActionConfig, OracleNavActionConfig, ActionConfig
from habitat.core.env import Env

from habitat.utils.humanoid_utils import MotionConverterSMPLX
from habitat.tasks.rearrange.actions.articulated_agent_action import ArticulatedAgentAction
from habitat.core.registry import registry
from gym import spaces

from habitat.datasets.rearrange.rearrange_dataset import RearrangeEpisode
import gzip
import json
import pandas as pd
from openpyxl.utils import get_column_letter
import copy
import random
import torch
import pathlib
import time
from collections import Counter

import cv2
from torchvision.transforms import Compose, Lambda
from torchvision.transforms.functional import to_pil_image
from torchvision.transforms._transforms_video import (
    CenterCropVideo,
    NormalizeVideo,
)
from pytorchvideo.data.encoded_video import EncodedVideo
from pytorchvideo.transforms import (
    ApplyTransformToKey,
    ShortSideScale,
    UniformTemporalSubsample,
    UniformCropVideo
) 
from typing import Dict
import time, datetime
import shutil

import git, os, gc
import glob
repo = git.Repo(".", search_parent_directories=True)
dir_path = repo.working_tree_dir
data_path = os.path.join(dir_path, "data")
os.chdir(dir_path)

from habitat.gpt.prompts.judge.prompt_intention_approval import approve_intention
from habitat.gpt.prompts.judge.prompt_predicate_approval import approve_predicate
from habitat.gpt.prompts.judge.prompt_category_approval import approve_category
from habitat.gpt.prompts.utils import load_response

from sentence_transformers import SentenceTransformer, util

from transformers import (AutoTokenizer,
                          AutoModelForSequenceClassification,
                          AutoModelForCausalLM,
                          BitsAndBytesConfig, 
                          Trainer, 
                          TrainingArguments,
                          pipeline,
                          DataCollatorForLanguageModeling)
from datasets import Dataset, DatasetDict, load_dataset
from peft import (LoraConfig, 
                  PeftConfig, 
                  PeftModel, 
                  get_peft_model,
                  prepare_model_for_kbit_training,
                  PeftModelForSequenceClassification)
from sklearn.metrics import accuracy_score, f1_score
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.utils import resample
from scipy.stats import pearsonr
from accelerate import dispatch_model
import tempfile
from dotenv import load_dotenv
from openai import OpenAI, OpenAIError


# TODO: Installing sentencepiece package causes "Segmentation fault (core dumped)" during model training. 
# Using 'mistralai/Mistral-7B-Instruct-v0.3' causes ValueError: Cannot instantiate this tokenizer from a slow version. If it's based on sentencepiece, make sure you have sentencepiece installed.
# Likely a package verison issue, should be solved
base_model_checkpoint_list = ['mistralai/Mistral-7B-Instruct-v0.2', 
                              'meta-llama/Meta-Llama-3.1-8B', 
                              'meta-llama/Meta-Llama-3.1-8B-Instruct',
                              "Qwen/Qwen2-7B-Instruct",
                              ]
base_model_checkpoint = base_model_checkpoint_list[0]


# Load environment variables from .env file
load_dotenv()
client = OpenAI(
  api_key=os.environ['OPENAI_API_KEY'],
)


def intention_approval_mllm(data_path, human_id, scene_id, time_tuple, intentions, temperature_dict, model_dict, method="main", collab=2, setting=1, start_over=False):
    day, file_idx, time_ = time_tuple
    output_dir = pathlib.Path(data_path) / "judge/gpt_response" / f"collaboration_{collab}/setting_{setting}" / f"{method}/intention_approval" / str(human_id).zfill(5) / scene_id / day
    os.makedirs(output_dir, exist_ok=True)
    conversation_hist = []

    if start_over:
        user, res = approve_intention(time_, intentions, output_dir, existing_response=None, temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, method=method, collab=collab)
        time.sleep(20)
    else:
        user, res = approve_intention(time_, intentions, output_dir, existing_response=load_response("intention_approval", output_dir, file_idx=file_idx), temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, method=method, collab=collab)
    conversation_hist.append([user, res])

    return conversation_hist


def predicate_approval_mllm(data_path, human_id, scene_id, time_tuple, human_thoughts, human_acts, robot_thoughts, robot_acts, temperature_dict, model_dict, method="main", collab=2, setting=1, start_over=False):
    day, file_idx, time_ = time_tuple
    output_dir = pathlib.Path(data_path) / "judge/gpt_response" / f"collaboration_{collab}/setting_{setting}" / f"{method}/predicate_approval" / str(human_id).zfill(5) / scene_id / day
    os.makedirs(output_dir, exist_ok=True)
    conversation_hist = []

    if start_over:
        user, res = approve_predicate(time_, human_thoughts, human_acts, robot_thoughts, robot_acts, output_dir, existing_response=None, temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, collab=collab)
        time.sleep(20)
    else:
        user, res = approve_predicate(time_, human_thoughts, human_acts, robot_thoughts, robot_acts, output_dir, existing_response=load_response("predicate_approval", output_dir, file_idx=file_idx), temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, collab=collab)
    conversation_hist.append([user, res])

    return conversation_hist


def category_approval_mllm(data_path, human_id, scene_id, time_tuple, human_thoughts, human_acts, robot_thoughts, robot_acts, temperature_dict, model_dict, method="main", collab=2, setting=1, start_over=False):
    day, file_idx, time_ = time_tuple
    output_dir = pathlib.Path(data_path) / "judge/gpt_response" / f"collaboration_{collab}/setting_{setting}" / f"{method}/category_approval" / str(human_id).zfill(5) / scene_id / day
    os.makedirs(output_dir, exist_ok=True)
    conversation_hist = []

    if start_over:
        user, res = approve_category(time_, human_thoughts, human_acts, robot_thoughts, robot_acts, output_dir, existing_response=None, temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, collab=collab)
        time.sleep(20)
    else:
        user, res = approve_category(time_, human_thoughts, human_acts, robot_thoughts, robot_acts, output_dir, existing_response=load_response("category_approval", output_dir, file_idx=file_idx), temperature_dict=temperature_dict, model_dict=model_dict, conversation_hist=conversation_hist, collab=collab)
    conversation_hist.append([user, res])

    return conversation_hist


def calculate_ocean_pearson_correlation(ocean1, ocean2_list, latest=True):
    """
    Calculate the Pearson Correlation between a ground truth OCEAN matrix and a list of OCEAN matrices.

    :param ocean1: Dictionary with OCEAN traits as keys and their corresponding scores as values (ground truth).
    :param ocean2_list: List of dictionaries where each dictionary has OCEAN traits as keys and their corresponding scores as values.
    :return: The Pearson Correlation between the ground truth OCEAN matrix and the majority-voted OCEAN matrix.
    """
    
    def round_to_nearest_half(value):
        """
        Round a value to the nearest 0.5 increment.
        """
        return round(value * 2) / 2
    
    # Convert ocean1 to a vector
    ocean1_vector = np.array([ocean1[trait] for trait in ocean1])

    # If latest is True, use only the last element of ocean2_list without rounding
    if latest:
        ocean2 = ocean2_list[-1]
        ocean2_vector = np.array([ocean2[trait] for trait in ocean2])
        # Calculate Pearson Correlation
        pearson_corr, _ = pearsonr(ocean1_vector, ocean2_vector)
        return ocean2, pearson_corr

    # Initialize the majority-voted OCEAN dictionary
    majority_voted_ocean = {}

    # If there is only one dictionary in the ocean2_list, take that as the majority voted result
    if len(ocean2_list) == 1:
        ocean2_list = [{k: round_to_nearest_half(v) for k, v in ocean2_list[0].items()}]

    # Iterate over each trait in the OCEAN model
    for trait in ocean1:
        # Get all the rounded values for this trait from each dictionary in ocean2_list
        rounded_values = [round_to_nearest_half(ocean[trait]) for ocean in ocean2_list]
        
        # Take the majority vote for this trait
        majority_vote = Counter(rounded_values).most_common(1)[0][0]
        majority_voted_ocean[trait] = majority_vote

    # Convert the majority-voted OCEAN dictionary to a vector
    majority_voted_ocean_vector = np.array([majority_voted_ocean[trait] for trait in majority_voted_ocean])

    # Calculate Pearson Correlation between ocean1 and the majority-voted ocean
    pearson_corr, _ = pearsonr(ocean1_vector, majority_voted_ocean_vector)
    
    return majority_voted_ocean, pearson_corr


def calculate_accuracy_and_f1(confidences, results, num_labels):
    """
    Parses LLM-generated responses to extract 'Yes' or 'No' answers, converts them to binary labels,
    and computes the accuracy and F1 scores.

    Parameters:
    confidences (list of str): A list of LLM-generated responses (answers).
    results (list of str): The true labels, where "Yes" or "No" are expected.

    Returns:
    accuracy (float): Accuracy score.
    f1_macro (float): Macro F1 score.
    f1_weighted (float): Weighted F1 score.
    """
    # Convert LLM responses to binary predictions (1 for 'Yes', 0 for 'No')
    predictions = [1 if 'yes' in response.lower() else 0 for response in confidences]
    
    # Convert true results ("Yes" -> 1, "No" -> 0)
    binary_results = [1 if 'yes' in result.lower() else 0 for result in results]

    # Calculate accuracy
    yes_indices = [i for i, pred in enumerate(predictions) if pred == 1]
    if yes_indices:
        filtered_predictions = [predictions[i] for i in yes_indices]
        filtered_results = [binary_results[i] for i in yes_indices]
        accuracy = accuracy_score(filtered_results, filtered_predictions)
    else:
        accuracy = 0  # If LLM never said "Yes", return 0 for accuracy
    # accuracy = accuracy_score(binary_results, predictions)
    
    # Calculate F1 scores
    f1_macro = f1_score(binary_results, predictions, average='macro')
    f1_weighted = f1_score(binary_results, predictions, average='weighted')

    if num_labels == 2:
        f1_binary = f1_score(binary_results, predictions, average='binary', labels=range(num_labels), pos_label=1)
        return accuracy, f1_macro, f1_weighted, f1_binary

    return accuracy, f1_macro, f1_weighted


def calculate_semantic_similarity(required_objects, predicted_objects, model="text-embedding-ada-002"):
    """
    Calculate average semantic similarity between required and predicted objects.
    For each required object, find its best match in predicted objects and average.
    
    Args:
        required_objects (list): List of required object names
        predicted_objects (list): List of predicted object names
    
    Returns:
        float: Average similarity score between 0-1
    """
    if len(predicted_objects) == 0:
        return 0.0
        
    # Get embeddings for all objects
    all_objects = required_objects + predicted_objects
    response = client.embeddings.create(
        model=model,
        input=all_objects
    )
    
    # Extract embeddings
    embeddings = [item.embedding for item in response.data]
    
    # Split embeddings back
    required_emb = np.array(embeddings[:len(required_objects)])
    predicted_emb = np.array(embeddings[len(required_objects):])
    
    # Compute similarity matrix
    cos_sims = cosine_similarity(required_emb, predicted_emb)
    
    # For each required object, get its best match score
    best_scores = np.max(cos_sims, axis=1)  # Best match for each required object
    
    # Return average of best scores
    return float(np.mean(best_scores))


# https://magazine.sebastianraschka.com/p/practical-tips-for-finetuning-llms
def select_model(data_type="intention", checkpoint_dir=None, pretrained=True):
    peft_config = LoraConfig(
        lora_alpha=16,
        lora_dropout=0.2,  # 0.1
        r=8,  # 4
        bias='none',
        task_type='CAUSAL_LM',
        target_modules=['q_proj', 'v_proj', 'k_proj', 'o_proj']  # ['q_proj', 'v_proj', 'k_proj', 'o_proj', 'fc_in', 'fc_out', 'lm_head']
    )

    # bnb_config = BitsAndBytesConfig(
    #     load_in_4bit = True,
    #     bnb_4bit_quant_type = 'nf4',
    #     bnb_4bit_compute_dtype = torch.bfloat16,
    #     bnb_4bit_use_double_quant = True,
    # )

    # bnb_config = BitsAndBytesConfig( 
    #     load_in_16bit=True,
    #     bnb_16bit_use_double_quant=False, 
    #     bnb_16bit_compute_dtype=torch.float16,
    #     bnb_16bit_quant_type="fp16"
    # )

    if pretrained and checkpoint_dir is not None:
        model = AutoModelForCausalLM.from_pretrained(
            checkpoint_dir,
            # quantization_config=bnb_config,
            # torch_dtype=torch.bfloat16,
            device_map='balanced',
            # low_cpu_mem_usage=True,  
            # max_memory = {
            #     0: '22GB',
            #     1: '23GB',
            #     2: '12GB',
            # }
        )
        model = PeftModel.from_pretrained(model, checkpoint_dir)
        # model = model.merge_and_unload()

    else:
        model = AutoModelForCausalLM.from_pretrained(
            base_model_checkpoint,
            use_auth_token=os.environ['HUGGINGFACE_TOKEN'],  
            # quantization_config=bnb_config,
            # torch_dtype=torch.bfloat16,
            device_map='balanced',
            # low_cpu_mem_usage=True,  # Ensure this is set to True for 4-bit/8-bit models
            trust_remote_code=True,
            # max_memory = {
            #     0: '22GB',
            #     1: '23GB',
            #     2: '12GB',
            # }
        )
        # Apply LoRA Configuration
        # model = prepare_model_for_kbit_training(model)
        model = get_peft_model(model, peft_config)
            
    return model


def create_tokenizer(model_checkpoint):
    # Load & Tokenize Data
    tokenizer = AutoTokenizer.from_pretrained(model_checkpoint)
    # tokenizer = AutoTokenizer.from_pretrained(
    #     model_checkpoint,
    #     use_auth_token=os.environ['HUGGINGFACE_TOKEN'],
    #     trust_remote_code=True
    # )
    tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "left"  # Pad on the left for causal LM
    
    return tokenizer


def clear_gpu_memory(model, trainer, tokenizer):
    # Step 1: Move any remaining models/tensors to CPU
    torch.cuda.empty_cache()

    # Step 2: Force garbage collection to clear up memory
    gc.collect()

    # Step 3: If using a model, explicitly move to CPU and delete
    # For example, if your model is still in memory:
    model.cpu()
    del model
    if trainer is not None: del trainer
    del tokenizer

    # Step 4: Clear the GPU cache
    torch.cuda.empty_cache()

    # Step 5: (Optional) Reset device states
    for device in range(torch.cuda.device_count()):
        with torch.cuda.device(device):
            torch.cuda.reset_max_memory_allocated()
            torch.cuda.reset_max_memory_cached()


def balance_dataset(data):
    # Separate majority and minority classes
    data_yes = [d for d in data if d['label'].lower() == 'yes']
    data_no = [d for d in data if d['label'].lower() == 'no']
    
    # Oversample minority class
    data_yes_oversampled = resample(data_yes, 
                                    replace=True,     # Sample with replacement
                                    n_samples=len(data_no),  # Match number in majority class
                                    random_state=42)  # Reproducible results
    
    # Combine majority class with oversampled minority class
    balanced_data = data_no + data_yes_oversampled
    
    # Shuffle the combined dataset
    random.shuffle(balanced_data)
    
    return balanced_data


# https://huggingface.co/docs/transformers/main/en/perf_train_gpu_one#mixed-precision-training
def train_model(epoch, data_train, data_test, time_tuple, output_path, data_type="intention", checkpoint_dir=None, pretrained=True):
    day, time_ =  time_tuple
    output_dir = str(pathlib.Path(output_path).parent)
    ts = time.time()
    time_string = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d-%H-%M-%S')
    output_time_path = os.path.join(f"{output_dir}/{day}_hist", f"{time_string}_{time_}")
    os.makedirs(output_path, exist_ok=True)
    # os.makedirs(output_time_path, exist_ok=True)

    model = select_model(data_type=data_type, checkpoint_dir=checkpoint_dir, pretrained=pretrained)
    # model.train()
    tokenizer = create_tokenizer(checkpoint_dir) if pretrained else create_tokenizer(base_model_checkpoint)

    data_collator = DataCollatorForLanguageModeling(
        tokenizer=tokenizer,
        mlm=False,
        pad_to_multiple_of=8
    )

    def tokenize_function(examples):
        # Combine the prompt and label into a single text sequence
        full_texts = [text + label for text, label in zip(examples['text'], examples['label'])]
        # Tokenize the combined texts
        model_inputs = tokenizer(full_texts, truncation=True, padding=True)
        # For causal language modeling, labels are the same as input_ids
        model_inputs["labels"] = model_inputs["input_ids"].copy()
        return model_inputs

    # Create datasets
    data_train = balance_dataset(data_train)
    data_test = balance_dataset(data_test)
    train_ds = Dataset.from_list(data_train)
    test_ds = Dataset.from_list(data_test)

    # Remove 'text' and 'label' columns after tokenization
    tokenized_train_ds = train_ds.map(tokenize_function, batched=True, remove_columns=['text', 'label'])
    tokenized_test_ds = test_ds.map(tokenize_function, batched=True, remove_columns=['text', 'label'])

    # Set format for PyTorch
    tokenized_train_ds.set_format(type='torch')
    tokenized_test_ds.set_format(type='torch')

    # Fine-tune LLM 
    training_args = TrainingArguments(
        output_dir=output_dir,
        learning_rate=1e-5,  # 5e-5
        per_device_train_batch_size=1,
        num_train_epochs=epoch,
        weight_decay=0.01,
        save_strategy="epoch",
        save_steps=1,
        save_total_limit=1,
        gradient_accumulation_steps=4,
        # gradient_checkpointing=True,
        # torch_empty_cache_steps=10,  # causes model training stuck
    )

    # if training_args.gradient_checkpointing:
    #     model.gradient_checkpointing_enable()

    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=tokenized_train_ds,
        tokenizer=tokenizer,
        # data_collator=data_collator,
    )
    trainer.train()

    # Save the trained model and tokenizer
    model.save_pretrained(output_path)
    tokenizer.save_pretrained(output_path)

    # Save additional output in the timestamped directory
    # model.save_pretrained(output_time_path)
    # tokenizer.save_pretrained(output_time_path)
    
    checkpoints = [d for d in os.listdir(output_dir) if d.startswith("checkpoint")]
    for checkpoint in checkpoints:
        full_checkpoint_path = os.path.join(output_dir, checkpoint)
        if os.path.exists(full_checkpoint_path):
            shutil.rmtree(full_checkpoint_path)

    # Free GPU memory
    clear_gpu_memory(model, trainer, tokenizer)


def test_model(data_test, checkpoint_dir, data_type="intention", cls_type="traits", pretrained=True):
    model = select_model(data_type=data_type, checkpoint_dir=checkpoint_dir, pretrained=pretrained)
    tokenizer = create_tokenizer(checkpoint_dir) if pretrained else create_tokenizer(base_model_checkpoint)

    pipe = pipeline(
        "text-generation", 
        model=model, 
        tokenizer = tokenizer, 
        torch_dtype=torch.float32,
        device_map="auto",
    )
    
    generated_answers = []
    with torch.no_grad():
        for data in data_test:
            prompt = data['text']
            # inputs = tokenizer(prompt, return_tensors="pt", truncation=True, padding=True)
            # # outputs = model.generate(
            # #     input_ids=inputs['input_ids'],
            # #     attention_mask=inputs['attention_mask'],
            # #     max_length=inputs['input_ids'].shape[1] + 10,  # Adjust max_length as needed
            # #     temperature=0.3,
            # #     top_p=0.9,
            # #     do_sample=True,
            # #     num_return_sequences=1,
            # #     eos_token_id=tokenizer.eos_token_id,
            # # )
            # outputs = model.generate(inputs.input_ids,
            #                         attention_mask=inputs['attention_mask'],
            #                         max_new_tokens=512
            #                         )

            # generated_text = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)
            
            sequences = pipe(
                prompt,
                do_sample=True,
                max_new_tokens=100, 
                temperature=0.2, 
                top_k=50, 
                top_p=0.95,
                num_return_sequences=3,
            )

            # Collect the generated answers
            answers = []
            for seq in sequences:
                generated_text = seq['generated_text']
                # Remove the prompt to get only the answer
                answer = generated_text[len(prompt):].strip()
                answers.append(answer)
            
            # Count how many answers contain "yes" and "no" (case-insensitive)
            yes_count = sum(1 for ans in answers if 'yes' in ans.lower())
            no_count = len(answers) - yes_count
            
            # Determine the majority answer
            majority_answer = 'Yes' if yes_count > no_count else 'No'
            generated_answers.append(majority_answer)

            print(f"Yes Count: {yes_count}, No Count: {no_count}")
            print(f"Majority Answer: {majority_answer}\n")

    print()
    print(f"Checkpoint: {checkpoint_dir}")
    print(f"Answers: {generated_answers}")
    print()

    clear_gpu_memory(model, None, tokenizer)
    return generated_answers


def create_data(texts, labels, time_, traits, hist, data_type="intention", collab=2):
    intentions_hist, predicates_hist = hist[0], hist[1]
    data_train = []

    if data_type == "intention":
        for intention, label in zip(texts, labels):
            if label is None:
                label = "No"
            prompt = (
                f"### Instruction:\n"
                f"Considering the human's profile, traits, temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{intention}'. Respond with 'Yes' or 'No'.\n\n"
                f"### Input:\n"
                f"Human Profile: {traits[0]}\n"
                f"Big Five Traits: {traits[1]}\n"
                f"Previous Relevant Intentions: {intentions_hist}\n"
                f"Current Time: {time_}\n\n"
                f"### Response:"
            )
            data_train.append({
                "text": prompt,
                "label": label
            })

    elif data_type == "predicates":
        _, thoughts, acts = texts[0], texts[1], texts[2]
        for thought, act, label in zip(thoughts, acts, labels):
            if label is None:
                label = "No"
            if collab == 1:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the human's profile, traits, temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Human Profile: {traits[0]}\n"
                    f"Big Five Traits: {traits[1]}\n"
                    f"Previous Relevant Intentions: {intentions_hist}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            elif collab == 2:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the human's profile, traits, temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}' while holding '{act}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Human Profile: {traits[0]}\n"
                    f"Big Five Traits: {traits[1]}\n"
                    f"Previous Relevant Intentions: {intentions_hist}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            data_train.append({
                "text": prompt,
                "label": label
            })

    data_test = data_train
    return data_train, data_test


def update_data_with_traits(data_train, data_test, new_traits):
    """Update dataset entries with new traits while preserving other fields."""
    
    def update_single_dataset(dataset):
        updated_data = []
        for item in dataset:
            # Get the original text and split into lines
            lines = item["text"].split("\n")
            updated_lines = []
            
            for line in lines:
                # Update Human Profile line
                if line.startswith("Human Profile:"):
                    updated_line = f"Human Profile: {new_traits[0]}"
                # Update Big Five Traits line
                elif line.startswith("Big Five Traits:"):
                    updated_line = f"Big Five Traits: {new_traits[1]}"
                else:
                    updated_line = line
                updated_lines.append(updated_line)
            
            # Rebuild the text with updated traits
            updated_text = "\n".join(updated_lines)
            
            updated_data.append({
                "text": updated_text,
                "label": item["label"]  # Preserve original label
            })
            
        return updated_data

    return update_single_dataset(data_train), update_single_dataset(data_test)


def create_data_no_traits(texts, labels, time_, traits, hist, data_type="intention", collab=2):
    intentions_hist, predicates_hist = hist[0], hist[1]
    data_train = []

    if data_type == "intention":
        for intention, label in zip(texts, labels):
            if label is None:
                label = "No"
            prompt = (
                f"### Instruction:\n"
                f"Considering the temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{intention}'. Respond with 'Yes' or 'No'.\n\n"
                f"### Input:\n"
                f"Previous Relevant Intentions: {intentions_hist}\n"
                f"Previous Relevant Tasks: {predicates_hist}\n"
                f"Current Time: {time_}\n\n"
                f"### Response:"
            )
            data_train.append({
                "text": prompt,
                "label": label
            })

    elif data_type == "predicates":
        _, thoughts, acts = texts[0], texts[1], texts[2]
        for thought, act, label in zip(thoughts, acts, labels):
            if label is None:
                label = "No"
            if collab == 1:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Previous Relevant Intentions: {intentions_hist}\n"
                    f"Previous Relevant Tasks: {predicates_hist}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            elif collab == 2:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the temporal dependence on past behaviors, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}' while holding '{act}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Previous Relevant Intentions: {intentions_hist}\n"
                    f"Previous Relevant Tasks: {predicates_hist}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            data_train.append({
                "text": prompt,
                "label": label
            })

    data_test = data_train
    return data_train, data_test


def create_data_no_dependence(texts, labels, time_, traits, hist, data_type="intention", collab=2):
    intentions_hist, predicates_hist = hist[0], hist[1]
    data_train = []

    if data_type == "intention":
        for intention, label in zip(texts, labels):
            if label is None:
                label = "No"
            prompt = (
                f"### Instruction:\n"
                f"Considering the human's profile, traits, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{intention}'. Respond with 'Yes' or 'No'.\n\n"
                f"### Input:\n"
                f"Human Profile: {traits[0]}\n"
                f"Big Five Traits: {traits[1]}\n"
                f"Current Time: {time_}\n\n"
                f"### Response:"
            )
            data_train.append({
                "text": prompt,
                "label": label
            })

    elif data_type == "predicates":
        _, thoughts, acts = texts[0], texts[1], texts[2]
        for thought, act, label in zip(thoughts, acts, labels):
            if label is None:
                label = "No"
            if collab == 1:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the human's profile, traits, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Human Profile: {traits[0]}\n"
                    f"Big Five Traits: {traits[1]}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            elif collab == 2:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the human's profile, traits, and the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}' while holding '{act}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Human Profile: {traits[0]}\n"
                    f"Big Five Traits: {traits[1]}\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            data_train.append({
                "text": prompt,
                "label": label
            })

    data_test = data_train
    return data_train, data_test


def create_data_none(texts, labels, time_, traits, hist, data_type="intention", collab=2):
    intentions_hist, predicates_hist = hist[0], hist[1]
    data_train = []

    if data_type == "intention":
        for intention, label in zip(texts, labels):
            if label is None:
                label = "No"
            prompt = (
                f"### Instruction:\n"
                f"Considering the current time: {time_}, determine if it is likely or unlikely that this human will: '{intention}'. Respond with 'Yes' or 'No'.\n\n"
                f"### Input:\n"
                f"Current Time: {time_}\n\n"
                f"### Response:"
            )
            data_train.append({
                "text": prompt,
                "label": label
            })

    elif data_type == "predicates":
        _, thoughts, acts = texts[0], texts[1], texts[2]
        for thought, act, label in zip(thoughts, acts, labels):
            if label is None:
                label = "No"
            if collab == 1:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            elif collab == 2:
                prompt = (
                    f"### Instruction:\n"
                    f"Considering the current time: {time_}, determine if it is likely or unlikely that this human will: '{thought}' while holding '{act}'. Respond with 'Yes' or 'No'.\n\n"
                    f"### Input:\n"
                    f"Current Time: {time_}\n\n"
                    f"### Response:"
                )
            data_train.append({
                "text": prompt,
                "label": label
            })

    data_test = data_train
    return data_train, data_test


def get_intention_idx(answer_intentions):
    yes_indices = [i for i, answer in enumerate(answer_intentions) if "yes" in answer.lower()]
    if yes_indices:
        return yes_indices  # Return multiple 'Yes'
    else:
        return [random.randint(0, len(answer_intentions) - 1)]  # Random index if no 'Yes'


def save_answers(answers, vis_dir, txt_path):
    # Ensure the directory exists
    os.makedirs(vis_dir, exist_ok=True)
    
    # Full path for the text file
    full_txt_path = os.path.join(vis_dir, txt_path)
    
    # Write the list to a text file, each answer on a new line
    with open(full_txt_path, 'w') as f:
        for answer in answers:
            f.write(f"{answer}\n")


def load_answers(vis_dir, txt_path):
    # Full path for the text file
    full_txt_path = os.path.join(vis_dir, txt_path)
    
    # Read the file and load it back into a list
    if os.path.exists(full_txt_path):
        with open(full_txt_path, 'r') as f:
            answers = [line.strip() for line in f.readlines()]
        return answers
    else:
        print(f"File not found: {full_txt_path}")
        return []


def save_results(eval_json_path, eval_intentions_llm_across_days, eval_predicates_llm_across_days, eval_predicates_semantic_across_days):
    """
    Save evaluation results with properly structured metrics.
    
    Each eval list contains tuples of:
    - For intentions/predicates LLM: (accuracy, f1_macro, f1_weighted, f1_binary)
    - For predicates semantic: (accuracy, f1_macro, f1_weighted, f1_binary, semantic_similarity)
    """
    # Extract metrics from the tuples
    acc_intentions = [day[0] for day in eval_intentions_llm_across_days]
    f1_macro_intentions = [day[1] for day in eval_intentions_llm_across_days]
    f1_weighted_intentions = [day[2] for day in eval_intentions_llm_across_days]
    f1_binary_intentions = [day[3] for day in eval_intentions_llm_across_days]
    # Compute average of the three F1 scores for each day
    f1_avg_intentions = [(day[1] + day[2] + day[3]) / 3 for day in eval_intentions_llm_across_days]
    
    acc_predicates_llm = [day[0] for day in eval_predicates_llm_across_days]
    f1_macro_predicates_llm = [day[1] for day in eval_predicates_llm_across_days]
    f1_weighted_predicates_llm = [day[2] for day in eval_predicates_llm_across_days]
    f1_binary_predicates_llm = [day[3] for day in eval_predicates_llm_across_days]
    # Compute average of the three F1 scores for each day
    f1_avg_predicates_llm = [(day[1] + day[2] + day[3]) / 3 for day in eval_predicates_llm_across_days]
    
    acc_predicates_category = [day[0] for day in eval_predicates_semantic_across_days]
    f1_macro_predicates_category = [day[1] for day in eval_predicates_semantic_across_days]
    f1_weighted_predicates_category = [day[2] for day in eval_predicates_semantic_across_days]
    f1_binary_predicates_category = [day[3] for day in eval_predicates_semantic_across_days]
    # Compute average of the three F1 scores for each day
    f1_avg_predicates_category = [(day[1] + day[2] + day[3]) / 3 for day in eval_predicates_semantic_across_days]
    semantic_similarity = [day[4] for day in eval_predicates_semantic_across_days]
    
    # Structure the data to save in JSON format
    data_to_save = {
        "acc_intentions": acc_intentions,
        "f1_macro_intentions": f1_macro_intentions,
        "f1_weighted_intentions": f1_weighted_intentions,
        "f1_binary_intentions": f1_binary_intentions,
        "f1_avg_intentions": f1_avg_intentions,
        
        "acc_predicates_llm": acc_predicates_llm,
        "f1_macro_predicates_llm": f1_macro_predicates_llm,
        "f1_weighted_predicates_llm": f1_weighted_predicates_llm,
        "f1_binary_predicates_llm": f1_binary_predicates_llm,
        "f1_avg_predicates_llm": f1_avg_predicates_llm,
        
        "acc_predicates_category": acc_predicates_category,
        "f1_macro_predicates_category": f1_macro_predicates_category,
        "f1_weighted_predicates_category": f1_weighted_predicates_category,
        "f1_binary_predicates_category": f1_binary_predicates_category,
        "f1_avg_predicates_category": f1_avg_predicates_category,
        "semantic_similarity": semantic_similarity
    }
    
    # Write data with each list on a new line for readability
    with open(eval_json_path, 'w') as f:
        f.write('{\n')
        for i, (key, value) in enumerate(data_to_save.items()):
            # Write each key-value pair on a new line, with the list on the same line
            json_line = f'    "{key}": {json.dumps(value)}'
            if i < len(data_to_save) - 1:
                json_line += ','  # Add a comma for all items except the last
            f.write(json_line + '\n')
        f.write('}\n')


def load_results(eval_json_path):
    """
    Load evaluation results from JSON file.
    
    Returns all metrics as separate lists.
    """
    # Read the data from the JSON file
    with open(eval_json_path, 'r') as f:
        data_loaded = json.load(f)
    
    # Extract each list and return them
    acc_intentions = data_loaded.get("acc_intentions", [])
    f1_macro_intentions = data_loaded.get("f1_macro_intentions", [])
    f1_weighted_intentions = data_loaded.get("f1_weighted_intentions", [])
    f1_binary_intentions = data_loaded.get("f1_binary_intentions", [])
    f1_avg_intentions = data_loaded.get("f1_avg_intentions", [])
    
    acc_predicates_llm = data_loaded.get("acc_predicates_llm", [])
    f1_macro_predicates_llm = data_loaded.get("f1_macro_predicates_llm", [])
    f1_weighted_predicates_llm = data_loaded.get("f1_weighted_predicates_llm", [])
    f1_binary_predicates_llm = data_loaded.get("f1_binary_predicates_llm", [])
    f1_avg_predicates_llm = data_loaded.get("f1_avg_predicates_llm", [])
    
    acc_predicates_category = data_loaded.get("acc_predicates_category", [])
    f1_macro_predicates_category = data_loaded.get("f1_macro_predicates_category", [])
    f1_weighted_predicates_category = data_loaded.get("f1_weighted_predicates_category", [])
    f1_binary_predicates_category = data_loaded.get("f1_binary_predicates_category", [])
    f1_avg_predicates_category = data_loaded.get("f1_avg_predicates_category", [])
    semantic_similarity = data_loaded.get("semantic_similarity", [])
    
    return (acc_intentions, f1_macro_intentions, f1_weighted_intentions, f1_binary_intentions, f1_avg_intentions,
            acc_predicates_llm, f1_macro_predicates_llm, f1_weighted_predicates_llm, f1_binary_predicates_llm, f1_avg_predicates_llm,
            acc_predicates_category, f1_macro_predicates_category, f1_weighted_predicates_category, f1_binary_predicates_category, f1_avg_predicates_category,
            semantic_similarity)


def append_evaluation_row(eval_csv_data, k, predicates_num, time_, gt_intention, 
                          pred_intentions, answer_intentions, human_thoughts, 
                          human_acts, robot_thoughts, robot_acts, answer_predicates,
                          profile_string, big_five, inferred_profile, inferred_traits,
                          eval_big_five, intentions_approval, eval_intentions,
                          predicates_approval, eval_predicates, category_approval, 
                          eval_semantic, method="main"):
    """Append a row to evaluation CSV data."""
    if method in ["main", "ag_human"]:
        if k == 0:
            # Full row
            eval_csv_data.append([time_, gt_intention, pred_intentions[k], 
                                answer_intentions[k], f"{human_thoughts[k]} {human_acts[k]}",
                                f"{robot_thoughts[k]} {robot_acts[k]}", answer_predicates[k],
                                profile_string, big_five, inferred_profile, inferred_traits,
                                eval_big_five, intentions_approval, eval_intentions,
                                predicates_approval, eval_predicates, category_approval, 
                                eval_semantic])
        elif k < predicates_num:
            # Partial row with human data
            eval_csv_data.append(["", "", pred_intentions[k], answer_intentions[k],
                                f"{human_thoughts[k]} {human_acts[k]}",
                                f"{robot_thoughts[k]} {robot_acts[k]}", answer_predicates[k]] 
                                + [""] * 11)
        else:
            # Partial row with only robot data
            eval_csv_data.append(["", "", "", "", "",
                                f"{robot_thoughts[k]} {robot_acts[k]}", answer_predicates[k]]
                                + [""] * 11)
    
    elif method in ["prompting", "oracle"]:
        # For prompting, we have single intention and direct tasks without classifiers
        for k in range(predicates_num):
            if k == 0:
                # Full row for first task
                eval_csv_data.append([
                    time_, 
                    gt_intention,  # Human ground truth intention
                    pred_intentions,  # Single robot predicted intention
                    f"{human_thoughts[k]} {human_acts[k]}",
                    f"{robot_thoughts[k]} {robot_acts[k]}", 
                    profile_string, 
                    big_five, 
                    inferred_profile, 
                    inferred_traits,
                    eval_big_five, 
                    intentions_approval, 
                    eval_intentions,
                    predicates_approval, 
                    eval_predicates, 
                    category_approval, 
                    eval_semantic
                ])
            else:
                # Partial rows for remaining tasks
                eval_csv_data.append([
                    "", "", "",
                    f"{human_thoughts[k]} {human_acts[k]}",
                    f"{robot_thoughts[k]} {robot_acts[k]}"
                ] + [""] * 11)


def save_evaluation_results(eval_csv_path, eval_txt_path, eval_csv_data, data_train_intentions, data_train_predicates, day, method="main"):
    """
    Save evaluation results to Excel and text files with proper formatting.
    
    Args:
        eval_csv_path: Path to the Excel file
        eval_txt_path: Path to the text file
        eval_csv_data: List of data rows for the CSV
        data_train_intentions: Training data for intentions
        data_train_predicates: Training data for predicates
        day: Current day string for sheet naming
    """
    # Define header for the Excel file
    if method in ["main", "ag_human"]:
        header = [
            "Time", 
            "Human Intention", 
            "Robot Inferred Intention", 
            "Robot Intention Classification", 
            "Human Task", 
            "Robot Inferred Task", 
            "Robot Task Classification", 
            "Human Traits", 
            "Human Big-5", 
            "Robot Inferred Traits", 
            "Robot Inferred Big-5", 
            "Big-5 Eval (Corr_latest, Corr_voting)", 
            "Intention LLM Approval", 
            "Intention LLM Eval (Acc, F1 (macro, weighted, binary))", 
            "Task LLM Approval", 
            "Task LLM Eval (Acc, F1 (macro, weighted, binary))", 
            "Task Object Category Approval", 
            "Task Object Category Eval (Acc, F1 (macro, weighted, binary)); Semantic Similarity"
        ]
    
        # Create DataFrame
        df = pd.DataFrame(eval_csv_data, columns=header)
        
        # Save to Excel with auto-width columns
        if not os.path.exists(eval_csv_path):
            # File doesn't exist, create new one
            with pd.ExcelWriter(eval_csv_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, index=False, sheet_name=f'eval_{day}')
                
                # Auto-adjust column widths
                worksheet = writer.sheets[f'eval_{day}']
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_length + 2
        else:
            # File exists, append with replace option
            with pd.ExcelWriter(eval_csv_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, index=False, sheet_name=f'eval_{day}')
                
                # Auto-adjust column widths
                worksheet = writer.sheets[f'eval_{day}']
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_length + 2
        
        # Save training data to text file
        with open(eval_txt_path, 'w') as f:
            f.write("data_train_intentions:\n")
            f.write(str(data_train_intentions) + "\n\n")
            f.write("data_train_predicates:\n")
            f.write(str(data_train_predicates) + "\n\n")
    
    elif method in ["prompting", "oracle"]:
        # Simplified header without classifier columns
        header = [
            "Time", 
            "Human Intention", 
            "Robot Inferred Intention",  # Single intention
            "Human Task", 
            "Robot Inferred Task", 
            "Human Traits", 
            "Human Big-5", 
            "Robot Inferred Traits", 
            "Robot Inferred Big-5", 
            "Big-5 Eval (Corr_latest, Corr_voting)", 
            "Intention LLM Approval",  
            "Intention Eval (Acc, F1 (macro, weighted, binary))",  
            "Task LLM Approval",  
            "Task Eval (Acc, F1 (macro, weighted, binary))", 
            "Task Object Category Approval", 
            "Task Object Category Eval; Semantic Similarity"
        ]
        
        # Create DataFrame
        df = pd.DataFrame(eval_csv_data, columns=header)
        
        # Save to Excel with auto-width columns
        if not os.path.exists(eval_csv_path):
            with pd.ExcelWriter(eval_csv_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, index=False, sheet_name=f'eval_{day}')
                worksheet = writer.sheets[f'eval_{day}']
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_length + 2
        else:
            with pd.ExcelWriter(eval_csv_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, index=False, sheet_name=f'eval_{day}')
                worksheet = writer.sheets[f'eval_{day}']
                for column in df:
                    column_length = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = column_length + 2
        
        # No training data for prompting (no classifiers to train)
        if eval_txt_path:
            with open(eval_txt_path, 'w') as f:
                f.write("Direct Prompting Method - No classifier training data\n")
                f.write(f"Day: {day}\n")
